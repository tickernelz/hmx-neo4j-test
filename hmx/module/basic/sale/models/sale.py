import io
import logging
import random
import time
from datetime import timedelta

from django.db import models
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from hmx import api
from hmx.exceptions import UserError
from hmx.tasks import generate_excel_report_task_template
from hmx.tools.celery import require_celery_worker, use_task
from hmx.tools.misc import profile


_logger = logging.getLogger(__name__)


class Sale(models.Model):
    """
    Sale Order model with Approval Workflow support.

    Uses Meta.inherit to inherit from ApprovalMixin which provides:
    - action_confirm, action_send_for_approval, action_approve, action_reject, etc.
    - Status choices configuration via _status_choices / _status_choices_add / _status

    To customize status flow, override:
    - _status_choices: Complete list of statuses (replaces base)
    - _status_choices_add: Additional statuses to add to base
    - _status_transitions: Define allowed status transitions
    - _on_confirm_no_workflow(): Called when no workflow is applicable

    Example custom flows:

    1. Simple flow (Draft → Approved):
        _status_choices = [
            ('draft', 'Draft'),
            ('approved', 'Approved'),
        ]
        _approval_pending_statuses = []
        _approval_approved_statuses = ['approved']

    2. Complex flow (Draft → Submitted → Under Review → Approved/Rejected → Completed):
        _status_choices = [
            ('draft', 'Draft'),
            ('submitted', 'Submitted'),
            ('under_review', 'Under Review'),
            ('revision_required', 'Revision Required'),
            ('approved', 'Approved'),
            ('rejected', 'Rejected'),
            ('completed', 'Completed'),
        ]
        _approval_pending_statuses = ['submitted', 'under_review']
        _approval_approved_statuses = ['approved', 'completed']
        _approval_rejected_statuses = ['rejected']
    """

    class Meta:
        # Must define name when using inherit to keep model's own identity
        name = 'sale'
        # Inherit from ApprovalMixin to get approval workflow functionality
        inherit = ['approvalmixin']

    # ===== STATUS CONFIGURATION =====
    # Default status choices from ApprovalMixin:
    # ('draft', 'Draft'), ('waiting_approval', 'Waiting For Approval'), ('approved', 'Approved'), ('rejected', 'Rejected')

    # To customize, uncomment and modify:
    # _status_choices = [
    #     ('draft', 'Draft'),
    #     ('submitted', 'Submitted'),
    #     ('under_review', 'Under Review'),
    #     ('approved', 'Approved'),
    #     ('rejected', 'Rejected'),
    #     ('completed', 'Completed'),
    # ]
    #
    # _status_transitions = {
    #     'draft': ['submitted'],
    #     'submitted': ['under_review', 'draft'],
    #     'under_review': ['approved', 'rejected', 'submitted'],
    #     'approved': ['completed', 'draft'],
    #     'rejected': ['draft'],
    #     'completed': [],
    # }
    #
    # _approval_pending_statuses = ['submitted', 'under_review']
    # _approval_approved_statuses = ['approved', 'completed']
    # _approval_rejected_statuses = ['rejected']

    @api.model
    def default_company(self):
        return self.env.company

    name = models.CharField(max_length=100, verbose_name=_("Order No"), null=True)
    company = models.ForeignKey('base.basecompany', on_delete=models.CASCADE, tracking=True, default=default_company)
    partner_id = models.ForeignKey(
        "partners.partner", verbose_name=_("Customer"), related_name="sale_ids", on_delete=models.CASCADE, null=True
    )
    quantity = models.FloatField(blank=True, null=True, verbose_name=_("Quantity"))
    product_ids = models.ManyToManyField(
        "product.products", verbose_name=_("Products"), related_name="sale_ids", blank=True
    )
    price = models.DecimalField(
        max_digits=10,
        decimal_ref='sale.product_price:digits_amount',
    )
    subtotal = models.FloatField(null=True, compute='_compute_subtotal', store=True, verbose_name=("Subtotal"))

    # Status field - choices will be set by Meta.inherit from ApprovalMixin
    # Or define your own choices here:
    status = models.CharField(
        max_length=50,
        choices=[
            ('draft', 'Draft'),
            ('waiting_approval', 'Waiting For Approval'),
            ('approved', 'Approved'),
            ('rejected', 'Rejected'),
        ],
        default='draft',
        verbose_name=_("Status"),
        tracking=True,
    )

    # Relation to approval workflow instance
    approval_workflow_instance = models.OneToOneField(
        'approval_workflow.approvalworkflowinstance',
        on_delete=models.SET_NULL,
        blank=True,
        null=True,
        verbose_name=_("Approval Workflow Instance"),
        related_name="sale_order",
    )

    document = models.FileField(upload_to='documents/', null=True, blank=True)
    date = models.DateField(_("Date"), null=True, blank=True)
    datetime = models.DateTimeField(_("Datetime"), null=True, blank=True)

    restricted_field = models.CharField(max_length=255, null=True)

    # Override _on_confirm_no_workflow if you want custom behavior when no workflow
    def _on_confirm_no_workflow(self):
        """Called when action_confirm is triggered but no workflow is applicable."""
        # Default: just approve directly
        self.status = 'approved'
        _logger.info(f"[Sale._on_confirm_no_workflow] {self.id} directly approved (no workflow)")

    @api.model
    def create(self, vals):
        # hotfix kecil: pastikan price ada (agar tidak melanggar NOT NULL DB)
        if 'price' not in vals or vals.get('price') is None:
            vals['price'] = 0  # atau Decimal('0') sesuai kebutuhan
        res = super(Sale, self).create(vals)
        base_sequence = self.env['basesequence'].next_by_code('sale.order')
        if base_sequence:
            res.name = base_sequence
        return res

    @api.depends('quantity', 'price')
    def _compute_subtotal(self):
        for record in self:
            record.subtotal = record.quantity * record.price

    def action_dummy(self):
        for _record in self:
            raise UserError('Sample raise error')
        return True

    def action_unlink(self):
        for record in self:
            record.unlink()
        return True

    def action_export_data(self):
        return self.action_export()

    @use_task(name='Generate 1M Records', fallback_to_sync=False)
    def action_generate_1m_records(self, log=None):
        """
        Generate one million sales order records as a background task.

        This method is decorated with @use_task, which means:
        1. It runs asynchronously (in background) using Celery
        2. The UI will show task progress during execution
        3. The task cannot be executed if Celery is unavailable (fallback_to_sync=False)

        Decorator parameters:
            name: The display name shown in the UI during task execution
            fallback_to_sync: If True, would execute synchronously when Celery is unavailable
                            (set to False here to prevent execution without Celery)

        Args:
            log: A callback function provided by the task system for progress reporting.
                Should be used as follows:
                - log(progress=25, text="Starting phase")  # Report 25% progress with message
                - log(progress=50, text="Halfway done")    # Report 50% progress with message
                - log(state="SUCCESS", progress=100, text="Complete")  # Mark as complete

        The log function will automatically update the UI with progress information.
        All methods decorated with @use_task should include the log parameter (default None).
        """

        if log:
            log(progress=0, text="Starting record generation process")

        Partner = self.env['partner']
        Product = self.env['products']
        user_id = self.env.user.id
        company_id = self.env.company.id
        now = timezone.now()

        cr = self._cr

        if log:
            log(progress=5, text="Creating partner records")

        partners = Partner.browse()
        for i in range(1, 11):
            partner_name = f"Partner {i}"
            partner = Partner.search([('name', '=', partner_name)], limit=1)
            if not partner:
                partner = Partner.create(
                    {
                        'name': partner_name,
                        'email': f'partner{i}@hmx.com',
                        'user_id': user_id,
                        'company': self.env.company.pk,
                        'branch': self.env.branch.pk,
                    }
                )

            partners += partner

        if log:
            log(progress=10, text="Creating product records")

        products = Product.browse()
        for i in range(1, 11):
            product_name = f"Product {i}"
            product = Product.search([('name', '=', product_name)], limit=1)
            if not product:
                product = Product.create(
                    {
                        'name': product_name,
                        'company': company_id,
                    }
                )
            products += product

        partner_ids = partners.ids
        product_ids = products.ids

        total_lines_target = 1_000_000
        lines_per_order = 10
        total_orders = total_lines_target // lines_per_order

        if log:
            log(progress=15, text=f"Starting to generate {total_orders:,} orders")

        order_buf = io.StringIO()
        for i in range(total_orders):
            partner_id = random.choice(partner_ids)
            order_buf.write(f"order_{i}\t{company_id}\t{partner_id}\t0.0\t{now}\t{now}\t{user_id}\t{user_id}\tdraft\n")
            if i % 50_000 == 0 and i > 0:
                progress = 15 + int((i / total_orders) * 40)
                if log:
                    log(progress=progress, text=f"Generated {i:,} orders ({i * 100 // total_orders}%)")

                order_buf.seek(0)
                cr.copy_from(
                    order_buf,
                    'sale_sale',
                    columns=[
                        'name',
                        'company_id',
                        'partner_id_id',
                        'price',
                        'created_at',
                        'updated_at',
                        'created_by',
                        'edited_by',
                        'status',
                    ],
                )
                order_buf = io.StringIO()

        order_buf.seek(0)
        cr.copy_from(
            order_buf,
            'sale_sale',
            columns=[
                'name',
                'company_id',
                'partner_id_id',
                'price',
                'created_at',
                'updated_at',
                'created_by',
                'edited_by',
                'status',
            ],
        )

        if log:
            log(progress=55, text="Orders created, fetching order IDs")

        cr.execute("SELECT id, name FROM sale_sale ORDER BY id DESC LIMIT %s", (total_orders,))
        order_ids = cr.fetchall()

        if log:
            log(progress=60, text="Starting to generate order lines")

        line_count = 0
        line_buf = io.StringIO()

        for idx, (order_id, order_name) in enumerate(order_ids):
            if idx % (total_orders // 20) == 0 and idx > 0:
                progress = 60 + int((idx / total_orders) * 35)
                if log:
                    log(
                        progress=progress,
                        text=f"Generated {line_count:,} lines ({idx * 100 // total_orders}% of orders)",
                    )

            for _line_idx in range(lines_per_order):
                product_id = random.choice(product_ids)
                qty = random.randint(1, 10)
                price = random.choice([10000, 20000, 30000, 40000, 50000])
                subtotal = qty * price
                line_buf.write(
                    f"{order_id}\t{order_name}\t{product_id}\t{qty}\t{price}\t{subtotal}\t{now}\t{now}\t{user_id}\t{user_id}\n"
                )

            line_count += lines_per_order

            if line_buf.tell() > 10_000_000:
                line_buf.seek(0)
                cr.copy_from(
                    line_buf,
                    'sale_saleorderline',
                    columns=[
                        'sale_id_id',
                        'name',
                        'product_id_id',
                        'quantity',
                        'price',
                        'subtotal',
                        'created_at',
                        'updated_at',
                        'created_by',
                        'edited_by',
                    ],
                )
                line_buf = io.StringIO()

        line_buf.seek(0)
        cr.copy_from(
            line_buf,
            'sale_saleorderline',
            columns=[
                'sale_id_id',
                'name',
                'product_id_id',
                'quantity',
                'price',
                'subtotal',
                'created_at',
                'updated_at',
                'created_by',
                'edited_by',
            ],
        )

        if log:
            log(state="SUCCESS", progress=100, text="Orders generated")

    @use_task(name='Generate Sample Data', fallback_to_sync=False)
    def action_generate_sample_data(self, log=None):  # noqa: C901
        """
        Generate one million sales order records with realistic patterns.

        Includes:
        - Seasonal variations (Q4 peaks, summer dips)
        - Year-over-year growth trend
        - Weekly patterns (weekday vs weekend)
        - Random fluctuations and occasional spikes
        """

        if log:
            log(progress=0, text="Starting record generation process")

        Partner = self.env['partner']
        Product = self.env['products']
        user_id = self.env.user.id
        company_id = self.env.company.id
        now = timezone.now()

        # Calculate date range: 10 years ago to now
        end_date = now
        start_date = now - timedelta(days=10 * 365)

        cr = self._cr

        if log:
            log(progress=5, text="Creating partner records")

        partners = Partner.browse()
        for i in range(1, 11):
            partner_name = f"Partner {i}"
            partner = Partner.search([('name', '=', partner_name)], limit=1)
            if not partner:
                partner = Partner.create(
                    {
                        'name': partner_name,
                        'email': f'partner{i}@hmx.com',
                        'user_id': user_id,
                        'company': self.env.company.pk,
                        'branch': self.env.branch.pk,
                    }
                )
            partners += partner

        if log:
            log(progress=10, text="Creating product records")

        products = Product.browse()
        for i in range(1, 11):
            product_name = f"Product {i}"
            product = Product.search([('name', '=', product_name)], limit=1)
            if not product:
                product = Product.create(
                    {
                        'name': product_name,
                        'company': company_id,
                    }
                )
            products += product

        partner_ids = partners.ids
        product_ids = products.ids

        total_lines_target = 1_000_000
        lines_per_order = 10
        total_orders = total_lines_target // lines_per_order

        if log:
            log(progress=15, text=f"Generating timestamps with realistic patterns for {total_orders:,} orders")

        # Generate realistic timestamps using probability-based sampling
        def get_date_weight(date, start_date):
            """Calculate probability weight for a given date"""
            years_elapsed = (date - start_date).days / 365.25

            # 1. Year-over-year growth (15% compound annual growth)
            growth_multiplier = 1.15**years_elapsed

            # 2. Seasonal pattern (strong Q4, weak summer)
            month = date.month
            if month in [11, 12]:  # November, December - holiday season
                seasonal_multiplier = 1.8
            elif month in [1]:  # January - post-holiday
                seasonal_multiplier = 1.3
            elif month in [6, 7, 8]:  # Summer slowdown
                seasonal_multiplier = 0.7
            elif month in [3, 4, 9, 10]:  # Spring and fall - moderate
                seasonal_multiplier = 1.0
            else:
                seasonal_multiplier = 0.9

            # 3. Weekly pattern (lower on weekends)
            weekday = date.weekday()
            if weekday >= 5:  # Saturday, Sunday
                weekly_multiplier = 0.4
            elif weekday == 0:  # Monday - busy
                weekly_multiplier = 1.2
            else:
                weekly_multiplier = 1.0

            return growth_multiplier * seasonal_multiplier * weekly_multiplier

        # Pre-calculate all days in the range
        total_days = (end_date - start_date).days
        all_dates = [start_date + timedelta(days=i) for i in range(total_days)]

        # Calculate weights for each day
        weights = [get_date_weight(date, start_date) for date in all_dates]

        # Generate timestamps by sampling from the weighted distribution
        timestamps = []
        for i in range(total_orders):
            # Sample a date based on weights
            selected_date = random.choices(all_dates, weights=weights, k=1)[0]

            # Add random time during business hours (8 AM to 6 PM)
            hour = random.randint(8, 18)
            minute = random.randint(0, 59)
            second = random.randint(0, 59)

            # 5% chance of occasional spike (outside business hours for variety)
            if random.random() < 0.05:
                hour = random.randint(0, 23)

            timestamp = selected_date.replace(hour=hour, minute=minute, second=second)
            timestamps.append(timestamp)

            if i % 10_000 == 0 and i > 0 and log:
                progress = 15 + int((i / total_orders) * 5)
                log(progress=progress, text=f"Generated {i:,}/{total_orders:,} timestamps")

        if log:
            log(progress=20, text="Timestamps generated, preparing order data with lines")

        # Pre-generate all order lines data to calculate totals
        orders_data = []
        for i in range(total_orders):
            partner_id = random.choice(partner_ids)
            created_at = timestamps[i]

            # Generate lines for this order
            order_lines = []
            total_price = 0
            total_quantity = 0

            for _i in range(lines_per_order):
                product_id = random.choice(product_ids)
                qty = random.randint(1, 10)
                price = random.choice([10000, 20000, 30000, 40000, 50000])
                subtotal = qty * price

                order_lines.append({'product_id': product_id, 'quantity': qty, 'price': price, 'subtotal': subtotal})

                total_price += subtotal
                total_quantity += qty

            orders_data.append(
                {
                    'name': f'order_{i}',
                    'partner_id': partner_id,
                    'created_at': created_at,
                    'total_price': total_price,
                    'total_quantity': total_quantity,
                    'lines': order_lines,
                }
            )

            if i % 10_000 == 0 and i > 0 and log:
                progress = 20 + int((i / total_orders) * 10)
                log(progress=progress, text=f"Prepared {i:,} orders with lines")

        if log:
            log(progress=30, text="Creating order records")

        # Insert orders with correct totals
        order_buf = io.StringIO()
        for i, order_data in enumerate(orders_data):
            order_buf.write(
                f"{order_data['name']}\t{company_id}\t{order_data['partner_id']}\t"
                f"{order_data['total_price']}\t{order_data['total_quantity']}\t"
                f"{order_data['created_at']}\t{now}\t{user_id}\t{user_id}\tdraft\n"
            )

            if i % 50_000 == 0 and i > 0:
                progress = 30 + int((i / total_orders) * 25)
                if log:
                    log(progress=progress, text=f"Inserted {i:,} orders ({i * 100 // total_orders}%)")

                order_buf.seek(0)
                cr.copy_from(
                    order_buf,
                    'sale_sale',
                    columns=[
                        'name',
                        'company_id',
                        'partner_id_id',
                        'price',
                        'quantity',
                        'created_at',
                        'updated_at',
                        'created_by',
                        'edited_by',
                        'status',
                    ],
                )
                order_buf = io.StringIO()

        order_buf.seek(0)
        cr.copy_from(
            order_buf,
            'sale_sale',
            columns=[
                'name',
                'company_id',
                'partner_id_id',
                'price',
                'quantity',
                'created_at',
                'updated_at',
                'created_by',
                'edited_by',
                'status',
            ],
        )

        if log:
            log(progress=55, text="Orders created, fetching order IDs")

        cr.execute("SELECT id, name FROM sale_sale ORDER BY id DESC LIMIT %s", (total_orders,))
        order_records = cr.fetchall()

        if log:
            log(progress=60, text="Starting to generate order lines")

        line_count = 0
        line_buf = io.StringIO()

        # Create a mapping of order names to IDs
        order_name_to_id = {name: id for id, name in order_records}

        for idx, order_data in enumerate(orders_data):
            if idx % (total_orders // 20) == 0 and idx > 0:
                progress = 60 + int((idx / total_orders) * 35)
                if log:
                    log(
                        progress=progress,
                        text=f"Generated {line_count:,} lines ({idx * 100 // total_orders}% of orders)",
                    )

            order_id = order_name_to_id.get(order_data['name'])
            if not order_id:
                continue

            for line_data in order_data['lines']:
                line_buf.write(
                    f"{order_id}\t{order_data['name']}\t{line_data['product_id']}\t"
                    f"{line_data['quantity']}\t{line_data['price']}\t{line_data['subtotal']}\t"
                    f"{order_data['created_at']}\t{now}\t{user_id}\t{user_id}\n"
                )

            line_count += lines_per_order

            if line_buf.tell() > 10_000_000:
                line_buf.seek(0)
                cr.copy_from(
                    line_buf,
                    'sale_saleorderline',
                    columns=[
                        'sale_id_id',
                        'name',
                        'product_id_id',
                        'quantity',
                        'price',
                        'subtotal',
                        'created_at',
                        'updated_at',
                        'created_by',
                        'edited_by',
                    ],
                )
                line_buf = io.StringIO()

        line_buf.seek(0)
        cr.copy_from(
            line_buf,
            'sale_saleorderline',
            columns=[
                'sale_id_id',
                'name',
                'product_id_id',
                'quantity',
                'price',
                'subtotal',
                'created_at',
                'updated_at',
                'created_by',
                'edited_by',
            ],
        )

        if log:
            log(state="SUCCESS", progress=100, text="Orders generated with realistic patterns")

    @profile
    def some_heavy_method(self):
        self.search_read([], limit=10_000)

    def to_excel_safe(self, value=None):
        """Ensure value written to OpenPyXL is a valid primitive."""
        if isinstance(value, (int, float, bool)):
            return value
        if value is None:
            return ""
        return str(value)

    def get_xlsx_report(self, context=None, stream=None):
        """
        Generate XLSX report (Odoo-style compatible).
        Supports localization, clean formatting & auto column fit.
        """
        context = context or {}
        stream = stream or io.BytesIO()

        wb = Workbook()
        sheet = wb.active
        sheet.title = context.get("sheet_title", "Sale Order")

        # ===== Styles =====
        style_title = Font(size=16, bold=True)
        style_header = Font(bold=True)
        style_total = Font(bold=True)

        border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )
        fill_header = PatternFill("solid", fgColor="EEEEEE")

        # ===== Title =====
        sheet.merge_cells("A1:D1")
        sheet["A1"].value = f"{self._meta.verbose_name or 'Order'} #{self.name}"
        sheet["A1"].font = style_title
        sheet["A1"].alignment = Alignment(horizontal="center")

        # ===== Partner =====
        sheet["A3"].value = "Customer:"
        sheet["A3"].font = style_header
        sheet["B3"].value = self.partner_id.name or "-"

        if getattr(self.partner_id, "email", False):
            sheet["A4"].value = "Email:"
            sheet["A4"].font = style_header
            sheet["B4"].value = self.partner_id.email or "-"

        # ===== Table Header =====
        columns = [
            str(_("Product")),
            str(_("Qty")),
            str(_("Price")),
            str(_("Subtotal")),
        ]

        row_index = 6
        for col_index, label in enumerate(columns, 1):
            cell = sheet.cell(row=row_index, column=col_index, value=label)
            cell.font = style_header
            cell.fill = fill_header
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        # ===== Table Rows =====
        row_index += 1
        for line in self.lines.all():
            values = [
                self.to_excel_safe(line.product_id.name),
                line.quantity,
                line.price,
                line.subtotal,
            ]

            for col_index, value in enumerate(values, 1):
                cell = sheet.cell(row=row_index, column=col_index, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left")

            row_index += 1

        # ===== Total Row =====
        total_row = row_index + 1
        sheet[f"A{total_row}"].value = str(_("Total"))
        sheet[f"A{total_row}"].font = style_total

        sheet[f"D{total_row}"].value = sum(self.lines.values_list("subtotal", flat=True))
        sheet[f"D{total_row}"].font = style_total
        sheet[f"D{total_row}"].alignment = Alignment(horizontal="right")

        # ===== Auto-fit Columns (Safe for MergedCells) =====
        for col_cells in sheet.columns:
            max_len = 0
            col_index = None

            for cell in col_cells:
                if isinstance(cell, MergedCell):
                    continue
                if col_index is None:
                    col_index = cell.column

                if cell.value:
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except (TypeError, AttributeError):
                        pass

            if col_index:
                sheet.column_dimensions[get_column_letter(col_index)].width = max(max_len + 2, 12)

        # ===== Save Stream =====
        wb.save(stream)
        stream.seek(0)
        return stream

    def action_excel_data(self):
        return self.action_export_xlsx()

    @require_celery_worker()
    def action_excel_template(self):
        """
        Submit Excel report to Celery using template JSON.
        """
        record = self.ensure_one()
        report_ref = 'sale.sale_report_excel'
        task = generate_excel_report_task_template.delay(
            model_name=self._meta.label, record_ids=[record.id], report_ref_xml=report_ref
        )
        return {'success': True, "name": "Export XLSX Template", "type": "download", 'task_id': task.id}

    def action_sum(self, n=1_000_000_000):
        start = time.time()
        res = 0
        for i in range(n):
            res += i
        finish = time.time() - start
        print("Result: %s. Finished in %.2fs" % (res, finish))
        return res

    def action_sum_cy(self, n=1_000_000_000):
        from sale.engine.ops import do_sum

        start = time.time()
        res = do_sum(n)
        finish = time.time() - start
        print("Result: %s. Finished in %.2fs" % (res, finish))
        return res


class SaleOrderLine(models.Model):
    sale_id = models.ForeignKey(
        "sale", verbose_name=_("Sale"), related_name="lines", on_delete=models.CASCADE, null=True
    )
    name = models.CharField(max_length=100, verbose_name=_("Order No"), null=True)
    product_id = models.ForeignKey(
        "product.products", verbose_name=_("Products"), related_name="orderline", on_delete=models.CASCADE, null=True
    )
    quantity = models.FloatField(blank=True, null=True, verbose_name=_("Quantity"))
    price = models.FloatField(blank=True, null=True, verbose_name=_("Price"))
    subtotal = models.FloatField(null=True, compute='_compute_subtotal', store=True, verbose_name=("Subtotal"))

    # Child nested: self-parent relationship
    parent_id = models.ForeignKey(
        'self', verbose_name=_("Parent Line"), on_delete=models.CASCADE, null=True, blank=True, related_name='child_ids'
    )

    @api.depends('quantity', 'price')
    def _compute_subtotal(self):
        for record in self:
            record.subtotal = record.quantity * record.price

    @api.model_create_multi
    def create(self, vals_list, **kwargs):
        """
        Override create to handle self-parent relationships with clientKey references.
        When child has parent_id set to a clientKey string, we resolve it after batch create.
        """
        import logging

        _logger = logging.getLogger(__name__)

        needs_parent_resolution = []

        for idx, vals in enumerate(vals_list):
            parent_id_value = vals.get('parent_id')
            if parent_id_value and isinstance(parent_id_value, str) and not parent_id_value.isdigit():
                needs_parent_resolution.append(
                    {
                        'index': idx,
                        'clientKey': parent_id_value,
                        'original_vals': vals,
                    }
                )
                _logger.info(f'[SaleOrderLine] Child at index {idx} has clientKey parent: {parent_id_value}')
                vals['parent_id'] = None

        records = super(SaleOrderLine, self).create(vals_list, **kwargs)

        if needs_parent_resolution:
            _logger.info(f'[SaleOrderLine] Resolving {len(needs_parent_resolution)} children with clientKey parents')
            records_list = list(records)

            for resolution_info in needs_parent_resolution:
                child_idx = resolution_info['index']
                parent_clientKey = resolution_info['clientKey']
                child_record = records_list[child_idx]

                parent_record = None
                for parent_idx in range(child_idx):
                    candidate = records_list[parent_idx]
                    if not candidate.parent_id:
                        parent_record = candidate

                if parent_record:
                    child_record.write({'parent_id': parent_record.id})
                    _logger.info(
                        f'[SaleOrderLine] Resolved parent_id for child {child_record.id}: '
                        f'parent={parent_record.id}, clientKey={parent_clientKey}'
                    )
                else:
                    _logger.warning(f'[SaleOrderLine] Could not find parent for child {child_record.id}')

        return records


class Criteria(models.Model):
    sale_id = models.ForeignKey(
        "sale", verbose_name=_("Sale"), related_name="criteria", on_delete=models.CASCADE, null=True
    )
    name = models.CharField(max_length=100, verbose_name=_("Criteria Name"), null=True)
    product_id = models.ForeignKey(
        "product.products",
        verbose_name=_("Products"),
        related_name="product_criteria",
        on_delete=models.CASCADE,
        null=True,
    )
    quantity = models.FloatField(blank=True, null=True, verbose_name=_("Quantity"))
